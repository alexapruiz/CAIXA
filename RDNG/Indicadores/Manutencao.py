import sys
import openpyxl
from datetime import datetime


# Importando os arquivos das funções comuns
sys.path.append('c:\\Projetos\\CAIXA\\Funcoes')
from MessageBox import MessageBox
from BancodeDados import SQLite


class Manutencao():

    def __init__(self, arquivo_DB):
        # Abre a conexao do banco de dados
        self.BancodeDados = SQLite(arquivo_DB)


    def ProcuraLinhaCabecalho(aba_planilha):
        try:
            for x in range(1,100):
                celula_atual = aba_planilha.cell(row = x, column = 1)
                if celula_atual.value == "Comunidade":
                    return x
            
        except Exception as erro:
            print(f"Ocorreu um erro durante o processo de identificacao do cabecalho da planilha - ProcuraLinhaCabecalho - [' {erro}' ]")
            return


    def ImportaPlanilha(self,arquivo):
        # Le a planilha e insere os dados na tabela
        try:
            planilha = openpyxl.load_workbook(arquivo)
            aba_planilha = planilha.active
            linha_cabecalho = Manutencao.ProcuraLinhaCabecalho(aba_planilha)
            for linha in range(linha_cabecalho+1,aba_planilha.max_row):
                # Define o valor dos indicadores
                if ( (str(aba_planilha.cell(linha,13).value).find('01-Requisitos') == -1) and (str(aba_planilha.cell(linha,13).value).upper().find('MIGRA') == -1) ):
                    INDICADOR_CAMINHO_CC_INVALIDO = '1'
                else:
                    INDICADOR_CAMINHO_CC_INVALIDO = '0'

                # JUSTIFICATIVA AUSENCIA REQ
                if ( (str(aba_planilha.cell(linha,11).value).upper().find('MIGRA') > -1) ):
                    INDICADOR_JUSTIFICATIVA = '1'
                else:
                    INDICADOR_JUSTIFICATIVA = '0'

                self.BancodeDados.ExecutaComandoSQL(f"INSERT INTO INDICADORES_RDNG (DATA_IMPORTACAO,COMUNIDADE,SQUAD,SISTEMA,QTDE_HU_TOTAL,HU,URL_HU,QTDE_RASTREADO,URL_REQUISITO,QTDE_NAO_RASTREADO,QTDE_JUSTIF_AUSENCIA_REQ,INDICADOR_JUSTIFICATIVA,JUSTIF_AUSENCIA_REQ,QTDE_CCRC,INCLUSAO_REQ_CLEARCASE,INDICADOR_CAMINHO_CC_INVALIDO,FSW) VALUES ('{datetime.now().strftime('%Y%m%d')}','{aba_planilha.cell(linha,1).value}','{aba_planilha.cell(linha,2).value}','{aba_planilha.cell(linha,3).value}','{aba_planilha.cell(linha,4).value}','{aba_planilha.cell(linha,5).value}','{aba_planilha.cell(linha,6).value}','{aba_planilha.cell(linha,7).value}','{aba_planilha.cell(linha,8).value}','{aba_planilha.cell(linha,9).value}','{aba_planilha.cell(linha,10).value}','{INDICADOR_JUSTIFICATIVA}','{aba_planilha.cell(linha,11).value}','{aba_planilha.cell(linha,12).value}','{aba_planilha.cell(linha,13).value}','{INDICADOR_CAMINHO_CC_INVALIDO}','{aba_planilha.cell(linha,14).value}')")

        except Exception as erro:
            print(f"Ocorreu um erro durante a importacao da planilha para a base de dados - Manutencao.ImportaPlanilha - [' {erro}' ]")
            return


    def RemoveDados(self,data_importacao):
        try:
            self.BancodeDados.ExecutaComandoSQL(f"DELETE FROM INDICADORES_RDNG WHERE DATA_IMPORTACAO = '{data_importacao}'")

        except Exception as erro:
            print(f"Ocorreu um erro durante a remoção de dados - Manutencao.RemoveDados - [' {erro}' ]")
            return