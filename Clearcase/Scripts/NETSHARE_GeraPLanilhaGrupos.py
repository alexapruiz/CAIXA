#Importando as bibliotecas
import os
import sys
from openpyxl import Workbook

sys.path.insert(0, 'c://projetos//CAIXA//Funcoes')
from BancodeDados import SQLite


def Grava_Dados_VOB_Comunidade_Fabrica(VOB,COMUNIDADE,FABRICA):
    # Atualiza as informações de Comunidade e Fábrica da VOB
    BancodeDados.ExecutaComandoSQL(f"UPDATE VOBs SET COMUNIDADE = '{COMUNIDADE}' , FABRICA = '{FABRICA}' WHERE VOB = '{VOB}'")


def InsereDadosCompartilhamento(NOME_COMPART,CAMINHO_COMPART,GRUPO_COMPART):
    # Grava o caminho do compartilhamento da tabela de VOBs
    BancodeDados.ExecutaComandoSQL(f"UPDATE VOBs SET CAMINHO_COMPART = '{CAMINHO_COMPART}' WHERE VOB = '{NOME_COMPART}'")

    # Insere registro com nome do compartilhamento e o grupo
    BancodeDados.ExecutaComandoSQL(f"INSERT INTO COMPARTILHAMENTOS (NOME_COMPART, GRUPO_COMPART) VALUES ('{NOME_COMPART}','{GRUPO_COMPART}')")


def LeDados_NETSHARE(arquivo):
    # Abre os arquivos com o resultado dos comandos 'NET SHARE' e grava na base de dados
    try:
        NOME_COMPART = ""
        CAMINHO_COMPART = ""
        GRUPO_COMPART = ""
        with open(caminho + f"\\Arquivos\\{arquivo}", encoding = 'UTF-8') as arquivo:
            for linha in arquivo:
                linha = linha.upper()
                linha = linha.strip()

                if linha.find("SHARE NAME") > -1:
                    NOME_COMPART = linha[18:-1]

                if linha.find("PATH") > -1:
                    CAMINHO_COMPART = linha[18:]

                if linha.find("CORPCAIXA") > -1:
                    GRUPO_COMPART = linha[int(linha.find("CORPCAIXA")+10):int(linha.find(","))]

                    InsereDadosCompartilhamento(NOME_COMPART,CAMINHO_COMPART,GRUPO_COMPART)

    except Exception as erro:
        print(f"Ocorreu um erro durante a analise do arquivo: '{arquivo}' - LeDados_NETSHARE - [' {erro}' ]")
        return


def Atualiza_Base():
    # Limpa as tabelas e le os dados do LSVOB com a VOB e o servidor, depois le o arquivo DESC e atualiza a COMUNIDADE / FABRICA
	print("Limpando as tabelas...")

	BancodeDados.ExecutaComandoSQL(f"DELETE FROM VOBS")
	BancodeDados.ExecutaComandoSQL(f"DELETE FROM VOB_COMUNIDADE")
	BancodeDados.ExecutaComandoSQL(f"DELETE FROM COMPARTILHAMENTOS")

    # Arquivo LSVOB
	VOB = ""
	POSICAO_INICIAL = ""
	POSICAO_FINAL = ""
	try:
		print("Analisando arquivo LSVOB.txt ....")
		with open(f"{caminho}\\Arquivos\\LSVOB.txt", encoding = 'UTF-8') as Arquivo_LSVOB:
			for linha in Arquivo_LSVOB:
				# Para cada linha do arquivo LSVOB.txt, cria um comando DESC
				linha = linha.upper()

				POSICAO_INICIAL = linha.find("\\")
				POSICAO_FINAL = linha.find("\\\\")

				VOB = linha[linha.find("\\")+1 : linha.find("\\\\")].strip()

				SERVIDOR = linha[POSICAO_FINAL + 2:POSICAO_FINAL + 34].strip()
				if SERVIDOR.find("CADSVAPRNT002") > -1:
					SERVIDOR = "SAO PAULO"
				if SERVIDOR.find("CBRSVAPRNT005") > -1:
					SERVIDOR = "BRASILIA"
				if SERVIDOR.find("CADSVAPRNT009") > -1:
					SERVIDOR = "RIO DE JANEIRO"
				if SERVIDOR.find("CBRSVAPRNT010") > -1:
					SERVIDOR = "CEPEM"
				if SERVIDOR.find("CBRSVAPRNT013") > -1:
					SERVIDOR = "LOTERIAS"

				BancodeDados.ExecutaComandoSQL(f"INSERT INTO VOBS (VOB,SERVIDOR) VALUES ('{VOB}','{SERVIDOR}')")

	except Exception as erro:
		print(f"Ocorreu um erro durante a geração e leitura do arquivo LSVOB - Btn_Atualiza_Base_click - [' {erro}' ]")
		return

    # DESC
	try:
		VOB = ""
		COMUNIDADE = ""
		FABRICA = ""
		print("Analisando arquivo DESC.txt ....")

		with open(f"{caminho}\\Arquivos\\DESC.txt", encoding = 'UTF-8') as arquivo:
			for linha in arquivo:
				linha = linha.upper()
				linha = linha.strip()
				if linha.find("VERSIONED") > -1:
					if VOB != '':
						Grava_Dados_VOB_Comunidade_Fabrica(VOB,COMUNIDADE,FABRICA)

					VOB = linha[24:-1]
					COMUNIDADE=''
					FABRICA=''
				else:
					if linha.find("G DF5222 CC_ARRECADACAO") > -1:
						COMUNIDADE += " ARRECADACAO "
					if linha.find("G DF5222 CC_CAMBIO") > -1:
						COMUNIDADE += " CAMBIO "
					if linha.find("G DF5222 CC_CANAIS_CLIENTES") > -1:
						COMUNIDADE += " CANAIS_CLIENTES "
					if linha.find("G DF5222 CC_CANAIS_INTERNO") > -1:
						COMUNIDADE += " CANAIS_INTERNO "
					if linha.find("G DF5222 CC_DADOS") > -1:
						COMUNIDADE += " DADOS "
					if linha.find("G DF5222 CC_CONTRATACOES") > -1:
						COMUNIDADE += " CONTRATACOES "
					if linha.find("G DF5222 CC_CONTROLADORIA") > -1:
						COMUNIDADE += " CONTROLADORIA "
					if linha.find("G DF5222 CC_CREDITO") > -1:
						COMUNIDADE += " CREDITO "
					if linha.find("G DF5222 CC_DEPOSITO") > -1:
						COMUNIDADE += " DEPOSITO "
					if linha.find("G DF5222 CC_ESTRUTURANTES_TI") > -1:
						COMUNIDADE += " ESTRUTURANTES_TI "
					if linha.find("G DF5222 CC_FINANCEIRO") > -1:
						COMUNIDADE += " FINANCEIRO "
					if linha.find("G DF5222 CC_FOMENTO_DJ") > -1:
						COMUNIDADE += " FOMENTO_DJ "
					if linha.find("G DF5222 CC_FUNDOS_GOVERNO") > -1:
						COMUNIDADE += " FUNDOS_GOVERNO "
					if linha.find("G DF5222 CC_HABITACAO") > -1:
						COMUNIDADE += " HABITACAO "
					if linha.find("G DF5222 CC_INSTITUCIONAL") > -1:
						COMUNIDADE += " INSTITUCIONAL "
					if linha.find("G DF5222 CC_LOTERIAS") > -1:
						COMUNIDADE += " LOTERIAS "
					if linha.find("G DF5222 CC_MEIOS_PAGAMENTO") > -1:
						COMUNIDADE += " MEIOS_PAGAMENTO "
					if linha.find("G DF5222 CC_OPERACOES_BANCARIAS") > -1:
						COMUNIDADE += " OPERACOES_BANCARIAS "
					if linha.find("G DF5222 CC_OPEN_BANKING") > -1:
						COMUNIDADE += " OPEN_BANKING "
					if linha.find("G DF5222 CC_PESSOAS") > -1:
						COMUNIDADE += " PESSOAS "
					if linha.find("G DF5222 CC_PROGRAMAS_SOCIAIS") > -1:
						COMUNIDADE += " PROGRAMAS_SOCIAIS "
					if linha.find("G DF5222 CC_RISCO") > -1:
						COMUNIDADE += " RISCO "
					if linha.find("G DF5222 CC_SEGURANCA") > -1:
						COMUNIDADE += " SEGURANCA "
					if linha.find("G DF5222 CC_CRM_CADASTRO") > -1:
						COMUNIDADE += " CRM_CADASTRO "

					if linha.find("G DF5222 CC_ATOS") > -1:
						FABRICA += " ATOS "
					if linha.find("G DF5222 CC_BENNER") > -1:
						FABRICA += " BENNER "
					if linha.find("G DF5222 CC_BRQ") > -1:
						FABRICA += " BRQ "
					if linha.find("G DF5222 CC_BRY") > -1:
						FABRICA += " BRY "
					if linha.find("G DF5222 CC_CAST") > -1:
						FABRICA += " CAST "
					if linha.find("G DF5222 CC_CPQD") > -1:
						FABRICA += " CPQD "
					if linha.find("G DF5222 CC_DATAINFO") > -1:
						FABRICA += " DATAINFO "
					if linha.find("G DF5222 CC_ESEC") > -1:
						FABRICA += " ESEC "
					if linha.find("G DF5222 CC_FIRST") > -1:
						FABRICA += " FIRST "
					if linha.find("G DF5222 CC_FOTON") > -1:
						FABRICA += " FOTON "
					if linha.find("G DF5222 CC_GLOBALWEB") > -1:
						FABRICA += " GLOBALWEB "
					if linha.find("G DF5222 CC_INDRA") > -1:
						FABRICA += " INDRA "
					if linha.find("G DF5222 CC_LATIN") > -1:
						FABRICA += " LATIN "
					if linha.find("G DF5222 CC_MAGNA") > -1:
						FABRICA += " MAGNA "
					if linha.find("G DF5222 CC_MAPS") > -1:
						FABRICA += " MAPS "
					if linha.find("G DF5222 CC_MURAH") > -1:
						FABRICA += " MURAH "
					if linha.find("G DF5222 CC_QINTESS") > -1:
						FABRICA += " QINTESS "
					if linha.find("G DF5222 CC_RESOURCE") > -1:
						FABRICA += " RESOURCE "
					if linha.find("G DF5222 CC_RJE") > -1:
						FABRICA += " RJE "
					if linha.find("G DF5222 CC_SENIOR") > -1:
						FABRICA += " SENIOR "
					if linha.find("G DF5222 CC_SONDA") > -1:
						FABRICA += " SONDA "
					if linha.find("G DF5222 CC_SPREAD") > -1:
						FABRICA += " SPREAD "
					if linha.find("G DF5222 CC_STEFANINI") > -1:
						FABRICA += " STEFANINI "
					if linha.find("G DF5222 CC_TIVIT") > -1:
						FABRICA += " TIVIT "
					if linha.find("G DF5222 CC_TTY") > -1:
						FABRICA += " TTY "
					if linha.find("G DF5222 CC_TREE") > -1:
						FABRICA += " TREE "
					if linha.find("G DF5222 CC_VERT_SAS") > -1:
						FABRICA += " VERT_SAS "

					if linha.find("CORPCAIXA\G CS7266") > -1:
						COMUNIDADE = " MODELO ANTIGO - SP "
						FABRICA = " MODELO ANTIGO - SP "
					if linha.find("CORPCAIXA\G DF7390") > -1:
						COMUNIDADE = " MODELO ANTIGO - BR "
						FABRICA = " MODELO ANTIGO - BR "
					if linha.find("CORPCAIXA\G RJ7265") > -1:
						COMUNIDADE = " MODELO ANTIGO - RJ "
						FABRICA = " MODELO ANTIGO - RJ "
					if linha.find("CORPCAIXA\G DF5088") > -1:
						COMUNIDADE = " GRUPOS DF5088 "
						FABRICA = " GRUPOS DF5088 "

	except Exception as erro:
		print(f"Ocorreu um erro durante a importação do arquivo DESC - Btn_Atualiza_Base_click - [' {erro}' ]")
		return

    # Arquivo NETSHARE_BR
	print("Analisando arquivo NETSHARE_BR.txt ....")
	LeDados_NETSHARE('NETSHARE_BR.txt')
	print("Analisando arquivo NETSHARE_SP.txt ....")
	LeDados_NETSHARE('NETSHARE_SP.txt')
	print("Analisando arquivo NETSHARE_RJ.txt ....")
	LeDados_NETSHARE('NETSHARE_RJ.txt')
	print("Analisando arquivo NETSHARE_CEPEM.txt ....")
	LeDados_NETSHARE('NETSHARE_CEPEM.txt')
	print("Analisando arquivo NETSHARE_LOTERIAS.txt ....")
	LeDados_NETSHARE('NETSHARE_LOTERIAS.txt')


def GeraPlanilhaGrupos_Excel():
	# Gera uma planilha com as informações dos grupos, coletadas pelo NET SHARE
	try:
		arquivo_saida = Workbook()

		cursos_comunidades = BancodeDados.ConsultaSQL("SELECT DISTINCT(comunidade) FROM VOBs where COMUNIDADE IS NOT NULL AND COMUNIDADE <> '' ORDER BY COMUNIDADE")
		COMUNIDADES = cursos_comunidades.fetchone()
		while COMUNIDADES:
			cursor_grupos = BancodeDados.ConsultaSQL(f"SELECT DISTINCT(grupo_compart) FROM compartilhamentos WHERE grupo_compart like '%{COMUNIDADES[0].strip()}%' ORDER BY grupo_compart")
			GRUPOS = cursor_grupos.fetchone()

			if GRUPOS:
				NOME_COMUNIDADE = COMUNIDADES[0][0:30].strip()
				plan = arquivo_saida.create_sheet(title=NOME_COMUNIDADE)
				linha = 1

			while GRUPOS:
				plan[f"A{linha}"] = f"{GRUPOS[0].strip()}"
				linha += 1
				GRUPOS = cursor_grupos.fetchone()

			COMUNIDADES = cursos_comunidades.fetchone()

		# Remover a primeira aba (Sheet)
		aba_sheet = arquivo_saida['Sheet']
		arquivo_saida.remove(aba_sheet)

		# Salva a planilha
		arquivo_saida.save(caminho + '\\Arquivos\\Grupos_Clearcase.xlsx')

	except Exception as erro:
		print(f"Ocorreu um erro durante a geração do arquivo XLS dos grupos - GeraPlanilhaGrupos_Excel - [' {erro}' ]")
		return


def GeraPlanilhaVOBs_Excel():
	# Gera uma planilha com as VOBs, organizadas por comunidade
	try:
		arquivo_saida = Workbook()

		cursos_comunidades = BancodeDados.ConsultaSQL("SELECT DISTINCT(comunidade) FROM VOBs where COMUNIDADE IS NOT NULL AND COMUNIDADE <> '' ORDER BY COMUNIDADE")
		COMUNIDADES = cursos_comunidades.fetchone()
		while COMUNIDADES:
			cursor_vobs = BancodeDados.ConsultaSQL(f"SELECT VOB FROM VOBs WHERE COMUNIDADE like '%{COMUNIDADES[0].strip()}%'")
			VOBs = cursor_vobs.fetchone()

			if VOBs:
				NOME_COMUNIDADE = COMUNIDADES[0][0:30].strip()
				plan = arquivo_saida.create_sheet(title=NOME_COMUNIDADE)
				linha = 1

			while VOBs:
				plan[f"A{linha}"] = f"{VOBs[0].strip()}"
				linha += 1
				VOBs = cursor_vobs.fetchone()

			COMUNIDADES = cursos_comunidades.fetchone()

		# Remover a primeira aba (Sheet)
		aba_sheet = arquivo_saida['Sheet']
		arquivo_saida.remove(aba_sheet)

		# Salva a planilha
		arquivo_saida.save(caminho + '\\Arquivos\\VOBs_Clearcase.xlsx')

	except Exception as erro:
		print(f"Ocorreu um erro durante a geração do arquivo XLS dos grupos - GeraPlanilhaGrupos_Excel - [' {erro}' ]")
		return


def GeraPlanilhaVOBs_Grupos_Excel():
	# Gera uma planilha com as VOBs e os grupos, organizada por comunidade
	try:
		arquivo_saida = Workbook()

		cursos_comunidades = BancodeDados.ConsultaSQL("SELECT DISTINCT(comunidade) FROM VOBS where COMUNIDADE IS NOT NULL AND COMUNIDADE <> '' ORDER BY COMUNIDADE")
		COMUNIDADES = cursos_comunidades.fetchone()
		while COMUNIDADES:
			# Para cada comunidade, grava os grupos
			cursor_grupos = BancodeDados.ConsultaSQL(f"SELECT DISTINCT(grupo_compart) FROM compartilhamentos WHERE grupo_compart like '%{COMUNIDADES[0].strip()}%' ORDER BY grupo_compart")
			GRUPOS = cursor_grupos.fetchone()

			linha = 1			
			if GRUPOS:
				NOME_COMUNIDADE = COMUNIDADES[0][0:30].strip()
				plan = arquivo_saida.create_sheet(title=NOME_COMUNIDADE)

				while GRUPOS:
					plan[f"A{linha}"] = f"{GRUPOS[0].strip()}"
					linha += 1
					GRUPOS = cursor_grupos.fetchone()

				# Para cada comunidade, grava as VOBs
				cursor_vobs = BancodeDados.ConsultaSQL(f"SELECT DISTINCT(COMPART.NOME_COMPART) AS VOB, VOBs.FABRICA AS FABRICA , VOBs.SERVIDOR FROM COMPARTILHAMENTOS COMPART,VOBs WHERE GRUPO_COMPART like '%{NOME_COMUNIDADE}%' AND COMPART.NOME_COMPART = VOBs.VOB ORDER BY COMPART.NOME_COMPART")
				VOBs = cursor_vobs.fetchone()

				if VOBs:
					# Pulando 2 linhas para iniciar a gravação das VOBs
					linha += 2

				while VOBs:
					if VOBs[1] is not None:
						plan[f"A{linha}"] = f"{VOBs[0].strip()}" # Nome da VOB
						plan[f"B{linha}"] = f"{VOBs[1].strip()}" # Fabrica
						plan[f"C{linha}"] = f"{VOBs[2].strip()}" # Servidor
						linha += 1
					VOBs = cursor_vobs.fetchone()

			COMUNIDADES = cursos_comunidades.fetchone()

		# Remover a primeira aba (Sheet)
		aba_sheet = arquivo_saida['Sheet']
		arquivo_saida.remove(aba_sheet)

		# Salva a planilha
		arquivo_saida.save(caminho + '\\Arquivos\\VOBs_Grupos_Clearcase.xlsx')

	except Exception as erro:
		print(f"Ocorreu um erro durante a geração do arquivo XLS dos grupos - GeraPlanilhaGrupos_Excel - [' {erro}' ]")
		return


# Verifica se esta sendo chamado por um interpretador python ou executavel
caminho = sys.path[1]

# Abre a conexao do banco de dados
BancodeDados = SQLite(f"{caminho}\\Clearcase.db")

# Verifica se o arquivo já existe, para removê-lo
if os.path.exists(caminho + "\\Arquivos\\Grupos_Clearcase.xlsx"):
	os.remove(caminho + "\\Arquivos\\Grupos_Clearcase.xlsx")

# Cria a pasta 'Arquivos', caso não exista
if not os.path.exists(caminho + "\\Arquivos"):
	os.makedirs(caminho + "\\Arquivos")

#Atualiza_Base()

#GeraPlanilhaGrupos_Excel()

#GeraPlanilhaVOBs_Excel()

# Gera a planilha com as VOBs e os grupos
GeraPlanilhaVOBs_Grupos_Excel()

print("O arquivo com as informações dos grupos foi gerado com sucesso!!!")